import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

# Cargar el archivo CSV
csv_file = 'z4_Cargo_pre_cotizaciones_mensuales.csv'
df_csv = pd.read_csv(csv_file).drop(columns=['Ord'])

# Cargar el archivo Excel existente
excel_file = 'c:/Users/baran/OneDrive/Público/CotizacionesMensuales.xlsx'
book = load_workbook(excel_file)

# Seleccionar la hoja activa
sheet = book.active

# Convertir los datos del DataFrame a una lista de listas
data = df_csv.values.tolist()

# Determinar la fila de inicio (14 en este caso) y la columna de inicio
start_row = 14
start_col = 1

# Obtener el formato de la última fila de la tabla
table = sheet.tables['T_Tabla']
end_cell = table.ref.split(':')[1]
end_row = sheet[end_cell].row
end_column = sheet[end_cell].column_letter

# Escribir los datos del CSV en la hoja de cálculo manteniendo el formato
for row_idx, row_data in enumerate(data, start=start_row):
    for col_idx, cell_value in enumerate(row_data, start=start_col):
        cell = sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        if row_idx >= end_row:
            source_cell = sheet.cell(row=end_row, column=col_idx)
            cell.fill = copy(source_cell.fill)
            cell.font = copy(source_cell.font)
            cell.border = copy(source_cell.border)
            cell.alignment = copy(source_cell.alignment)

# Actualizar la referencia de la tabla
new_end_row = end_row + len(data)
new_ref = f"{table.ref.split(':')[0]}:{end_column}{new_end_row}"
table.ref = new_ref

# Guardar el archivo Excel con los datos añadidos
book.save(excel_file)

# Mostrar mensaje de éxito
'Actualización de la tabla completada y formato preservado.'